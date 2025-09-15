import os
import logging
import threading
import pandas as pd
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import StringVar, filedialog, messagebox, ttk as tk_ttk
from tkinter.ttk import Combobox
from ttkbootstrap.scrolled import ScrolledText
from ttkbootstrap.tableview import Tableview
from database.banks_repository import get_all_banks
from database.bank_transaction_repository import get_transactions_by_bank
from database.pos_transactions_repository import get_transactions_by_bank as get_pos_transactions_by_bank
from database.repositories.accounting import get_transactions_by_bank
from database.reconciliation_results_repository import get_reconciliation_results
from config.settings import (
    DATA_DIR, DEFAULT_FONT, DEFAULT_FONT_SIZE,
    HEADER_FONT_SIZE, BUTTON_FONT_SIZE
)

# کلاس برای نمایش لاگ‌ها در UI
class UIHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        
    def emit(self, record):
        msg = self.format(record) + '\n'
        # ScrolledText در ttkbootstrap از state پشتیبانی نمی‌کند
        self.text_widget.insert('end', msg)
        self.text_widget.see('end')

class ReportTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.setup_logging()
        
        # متغیرهای مورد نیاز
        self.selected_bank_var = StringVar(value="همه موارد")
        self.selected_table_var = StringVar(value="بانک")
        self.selected_transaction_type_var = StringVar(value="همه موارد")
        self.selected_reconciliation_status_var = StringVar(value="همه موارد")
        self.status_var = StringVar(value="آماده برای ساخت گزارش...")
        
        # ایجاد ویجت‌ها
        self.create_widgets()
        self.load_banks_to_combobox()
        
        # داده‌های جدول
        self.data = []
        self.columns = []
        
    def setup_logging(self):
        """راه‌اندازی سیستم لاگینگ"""
        os.makedirs(DATA_DIR, exist_ok=True)
        
        # تنظیمات کلی لاگر
        self.logger = logging.getLogger('report.tab')
        self.logger.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        
        # لاگر برای فایل خطاها
        error_handler = logging.FileHandler(os.path.join(DATA_DIR, 'report_error.txt'), encoding='utf-8')
        error_handler.setLevel(logging.ERROR)
        error_handler.setFormatter(formatter)
        
        # لاگر برای فایل لاگ عمومی
        file_handler = logging.FileHandler(os.path.join(DATA_DIR, 'report_log.txt'), encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)
        
        # لاگر برای کنسول با پشتیبانی از UTF-8
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(error_handler)
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        # UI handler will be added after creating log_text widget

    def create_widgets(self):
        PADX = 8
        PADY = 8

        # تنظیم فونت‌ها - حذف bold برای ظاهر لایتتر
        self.default_font = (DEFAULT_FONT, DEFAULT_FONT_SIZE)
        self.header_font = (DEFAULT_FONT, HEADER_FONT_SIZE)
        self.button_font = (DEFAULT_FONT, BUTTON_FONT_SIZE)
        self.log_font = (DEFAULT_FONT, DEFAULT_FONT_SIZE - 1)  # کمی کوچکتر برای لاگ‌ها

        # تنظیم استایل‌ها - برای یکپارچگی و حل مشکل نمایش
        style = ttk.Style()
        style.configure('Header.TLabelframe', font=self.header_font)
        style.configure('Header.TLabelframe.Label', font=self.header_font)
        style.configure('Default.TLabel', font=self.default_font)
        style.configure('Default.TEntry', font=self.default_font)
        style.configure('Bold.TButton', font=self.button_font)
        
        # اطمینان از اعمال تغییرات
        self.update_idletasks()

        main_frame = ttk.Frame(self)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # === بخش تنظیمات گزارش ===
        settings_frame = ttk.LabelFrame(main_frame, text="تنظیمات گزارش", style='Header.TLabelframe')
        settings_frame.pack(fill="x", pady=5)
        
        # ردیف اول - انتخاب بانک و جدول
        row1_frame = ttk.Frame(settings_frame)
        row1_frame.pack(fill="x", padx=PADX, pady=PADY)
        
        ttk.Label(row1_frame, text="بانک:", style='Default.TLabel').pack(side="right", padx=PADX)
        bank_combobox = Combobox(row1_frame, textvariable=self.selected_bank_var, font=self.default_font, state="readonly")
        bank_combobox.pack(side="right", padx=PADX)
        
        ttk.Label(row1_frame, text="جدول:", style='Default.TLabel').pack(side="right", padx=PADX)
        table_combobox = Combobox(row1_frame, textvariable=self.selected_table_var, font=self.default_font, state="readonly")
        table_combobox['values'] = ("بانک", "حسابداری", "پوز", "نتایج مغایرت گیری")
        table_combobox.pack(side="right", padx=PADX)
        
        # ردیف دوم - نوع تراکنش و وضعیت رکورد
        row2_frame = ttk.Frame(settings_frame)
        row2_frame.pack(fill="x", padx=PADX, pady=PADY)
        
        ttk.Label(row2_frame, text="نوع تراکنش:", style='Default.TLabel').pack(side="right", padx=PADX)
        transaction_type_combobox = Combobox(row2_frame, textvariable=self.selected_transaction_type_var, font=self.default_font, state="readonly")
        # Map Persian to English values for database compatibility
        transaction_type_combobox['values'] = ("همه موارد", "دریافتی", "پرداختی", "پوز", "چک")
        transaction_type_combobox.pack(side="right", padx=PADX)
        
        ttk.Label(row2_frame, text="وضعیت رکورد:", style='Default.TLabel').pack(side="right", padx=PADX)
        status_combobox = Combobox(row2_frame, textvariable=self.selected_reconciliation_status_var, font=self.default_font, state="readonly")
        status_combobox['values'] = ("همه موارد", "مغایرت گیری شده", "مغایرت گیری نشده")
        status_combobox.pack(side="right", padx=PADX)
        
        # دکمه ساخت گزارش
        button_frame = ttk.Frame(settings_frame)
        button_frame.pack(fill="x", padx=PADX, pady=PADY)
        
        generate_report_button = ttk.Button(
            button_frame, text="ساخت گزارش", style='Bold.TButton',
            command=self.generate_report
        )
        generate_report_button.pack(side="left", padx=PADX)
        
        # === بخش نمایش داده‌ها ===
        data_frame = ttk.Frame(main_frame)
        data_frame.pack(fill="both", expand=True, pady=5)  # اجازه گسترش برای فضای کافی برای جدول
        
        # جدول نمایش داده‌ها
        self.table_frame = ttk.Frame(data_frame, height=500)  # افزایش ارتفاع برای فونت و سطرهای بزرگتر
        self.table_frame.pack(fill="both", expand=True)  # اجازه گسترش برای فضای بیشتر
        
        # وضعیت
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill="x", pady=5)
        
        ttk.Label(status_frame, textvariable=self.status_var, style='Default.TLabel').pack(side="right")
        
        # === بخش دکمه‌های عملیاتی ===
        actions_frame = ttk.Frame(main_frame)
        actions_frame.pack(fill="x", pady=5)
        
        export_excel_button = ttk.Button(
            actions_frame, text="صادر کردن به اکسل", style='Bold.TButton',
            command=self.export_to_excel
        )
        export_excel_button.pack(side="right", padx=PADX)
        
        print_report_button = ttk.Button(
            actions_frame, text="چاپ گزارش", style='Bold.TButton',
            command=self.print_report
        )
        print_report_button.pack(side="right", padx=PADX)
        
        export_pdf_button = ttk.Button(
            actions_frame, text="صدور به PDF", style='Bold.TButton',
            command=self.export_to_pdf
        )
        export_pdf_button.pack(side="right", padx=PADX)
        
        # بخش لاگ‌ها
        log_frame = ttk.LabelFrame(main_frame, text="گزارش عملیات", style='Header.TLabelframe')
        log_frame.pack(fill="x", pady=5)
        
        self.log_text = ScrolledText(log_frame, height=5, font=self.log_font)
        self.log_text.pack(fill="both", expand=True, padx=PADX, pady=PADY)
        
        # اضافه کردن UI handler به لاگر
        ui_handler = UIHandler(self.log_text)
        ui_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        ui_handler.setLevel(logging.INFO)
        self.logger.addHandler(ui_handler)
    
    def map_persian_transaction_type_to_english(self, persian_type):
        """تبدیل نوع تراکنش فارسی به انگلیسی برای واکشی دیتابیس"""
        mapping = {
            "همه موارد": None,  # None means no filter
            "دریافتی": "received",
            "پرداختی": "paid", 
            "پوز": "pos",
            "چک": "check"
        }
        return mapping.get(persian_type, None)
    
    def load_banks_to_combobox(self):
        """بارگذاری لیست بانک‌ها در کامبوباکس"""
        try:
            banks = get_all_banks()
            bank_names = ["همه موارد"] + [bank[1] for bank in banks]  # bank[1] is bank_name
            combobox = self.nametowidget(self.winfo_children()[0].winfo_children()[0].winfo_children()[0].winfo_children()[1])
            combobox['values'] = bank_names
            self.logger.info("لیست بانک‌ها با موفقیت بارگذاری شد")
        except Exception as e:
            self.logger.error(f"خطا در بارگذاری لیست بانک‌ها: {str(e)}")
    
    def generate_report(self):
        """ساخت گزارش بر اساس تنظیمات انتخاب شده"""
        try:
            self.logger.info("در حال ساخت گزارش...")
            self.status_var.set("در حال ساخت گزارش...")
            
            # پاک کردن جدول قبلی
            for widget in self.table_frame.winfo_children():
                widget.destroy()
            
            # دریافت پارامترهای گزارش
            selected_bank = self.selected_bank_var.get()
            selected_table = self.selected_table_var.get()
            selected_transaction_type_persian = self.selected_transaction_type_var.get()
            selected_transaction_type = self.map_persian_transaction_type_to_english(selected_transaction_type_persian)
            selected_reconciliation_status = self.selected_reconciliation_status_var.get()
            
            # تبدیل وضعیت مغایرت‌گیری به مقدار عددی
            is_reconciled = None
            if selected_reconciliation_status == "مغایرت گیری شده":
                is_reconciled = 1
            elif selected_reconciliation_status == "مغایرت گیری نشده":
                is_reconciled = 0
            
            # دریافت شناسه بانک
            bank_id = None
            if selected_bank != "همه موارد":
                banks = get_all_banks()
                for bank in banks:
                    if bank[1] == selected_bank:  # bank[1] is bank_name
                        bank_id = bank[0]  # bank[0] is bank_id
                        break
            
            # دریافت داده‌ها بر اساس جدول انتخاب شده
            self.data = []
            if selected_table == "بانک":
                self.get_bank_transactions(bank_id, selected_transaction_type, is_reconciled)
            elif selected_table == "حسابداری":
                self.get_accounting_transactions(bank_id, selected_transaction_type, is_reconciled)
            elif selected_table == "پوز":
                self.get_pos_transactions(bank_id, is_reconciled)
            elif selected_table == "نتایج مغایرت گیری":
                self.get_reconciliation_results_data(bank_id)
            
            # نمایش داده‌ها در جدول
            if self.data:
                self.display_data_in_table()
                self.status_var.set(f"تعداد {len(self.data)} رکورد یافت شد")
                self.logger.info(f"تعداد {len(self.data)} رکورد یافت شد")
            else:
                self.status_var.set("هیچ رکوردی یافت نشد")
                self.logger.info("هیچ رکوردی یافت نشد")
        except Exception as e:
            self.logger.error(f"خطا در ساخت گزارش: {str(e)}")
            self.status_var.set(f"خطا در ساخت گزارش: {str(e)}")
            messagebox.showerror("خطا", f"خطا در ساخت گزارش: {str(e)}")
    
    def get_bank_transactions(self, bank_id, transaction_type, is_reconciled):
        """دریافت تراکنش‌های بانکی با فیلترهای مشخص شده"""
        try:
            # ایجاد کوئری برای دریافت تراکنش‌های بانکی
            import sqlite3
            from config.settings import DB_PATH
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = "SELECT bt.*, b.bank_name FROM BankTransactions bt JOIN Banks b ON bt.bank_id = b.id WHERE 1=1"
            params = []
            
            if bank_id:
                query += " AND bt.bank_id = ?"
                params.append(bank_id)
            
            if transaction_type is not None:
                query += " AND bt.transaction_type = ?"
                params.append(transaction_type)
            
            if is_reconciled is not None:
                query += " AND bt.is_reconciled = ?"
                params.append(is_reconciled)
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # تبدیل نتایج به لیست دیکشنری
            self.data = [dict(row) for row in rows]
            self.columns = [
                {"text": "شناسه", "dataindex": "id"},
                {"text": "بانک", "dataindex": "bank_name"},
                {"text": "تاریخ", "dataindex": "transaction_date"},
                {"text": "زمان", "dataindex": "transaction_time"},
                {"text": "مبلغ", "dataindex": "amount"},
                {"text": "شرح", "dataindex": "description"},
                {"text": "شماره مرجع", "dataindex": "reference_number"},
                {"text": "شماره ترمینال", "dataindex": "extracted_terminal_id"},
                {"text": "شماره پیگیری", "dataindex": "extracted_tracking_number"},
                {"text": "نوع تراکنش", "dataindex": "transaction_type"},
                {"text": "شماره کارت", "dataindex": "source_card_number"},
                {"text": "مغایرت گیری شده", "dataindex": "is_reconciled"}
            ]
            
            conn.close()
        except Exception as e:
            self.logger.error(f"خطا در دریافت تراکنش‌های بانکی: {str(e)}")
            raise
    
    def get_accounting_transactions(self, bank_id, transaction_type, is_reconciled):
        """دریافت تراکنش‌های حسابداری با فیلترهای مشخص شده"""
        try:
            # ایجاد کوئری برای دریافت تراکنش‌های حسابداری
            import sqlite3
            from config.settings import DB_PATH
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = "SELECT at.*, b.bank_name FROM AccountingTransactions at JOIN Banks b ON at.bank_id = b.id WHERE 1=1"
            params = []
            
            if bank_id:
                query += " AND at.bank_id = ?"
                params.append(bank_id)
            
            if transaction_type is not None:
                query += " AND at.transaction_type = ?"
                params.append(transaction_type)
            
            if is_reconciled is not None:
                query += " AND at.is_reconciled = ?"
                params.append(is_reconciled)
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # تبدیل نتایج به لیست دیکشنری
            self.data = [dict(row) for row in rows]
            self.columns = [
                {"text": "شناسه", "dataindex": "id"},
                {"text": "بانک", "dataindex": "bank_name"},
                {"text": "نوع تراکنش", "dataindex": "transaction_type"},
                {"text": "شماره تراکنش", "dataindex": "transaction_number"},
                {"text": "مبلغ", "dataindex": "transaction_amount"},
                {"text": "تاریخ سررسید", "dataindex": "due_date"},
                {"text": "تاریخ وصول", "dataindex": "collection_date"},
                {"text": "نام مشتری", "dataindex": "customer_name"},
                {"text": "شرح", "dataindex": "description"},
                {"text": "مغایرت گیری شده", "dataindex": "is_reconciled"},
                {"text": "سیستم", "dataindex": "is_new_system"}
            ]
            
            conn.close()
        except Exception as e:
            self.logger.error(f"خطا در دریافت تراکنش‌های حسابداری: {str(e)}")
            raise
    
    def get_pos_transactions(self, bank_id, is_reconciled):
        """دریافت تراکنش‌های پوز با فیلترهای مشخص شده"""
        try:
            # ایجاد کوئری برای دریافت تراکنش‌های پوز
            import sqlite3
            from config.settings import DB_PATH
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = "SELECT pt.*, b.bank_name FROM PosTransactions pt JOIN Banks b ON pt.bank_id = b.id WHERE 1=1"
            params = []
            
            if bank_id:
                query += " AND pt.bank_id = ?"
                params.append(bank_id)
            
            if is_reconciled is not None:
                query += " AND pt.is_reconciled = ?"
                params.append(is_reconciled)
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # تبدیل نتایج به لیست دیکشنری
            self.data = [dict(row) for row in rows]
            self.columns = [
                {"text": "شناسه", "dataindex": "id"},
                {"text": "بانک", "dataindex": "bank_name"},
                {"text": "شماره ترمینال", "dataindex": "terminal_number"},
                {"text": "شماره پایانه", "dataindex": "terminal_id"},
                {"text": "شماره کارت", "dataindex": "card_number"},
                {"text": "تاریخ تراکنش", "dataindex": "transaction_date"},
                {"text": "مبلغ تراکنش", "dataindex": "transaction_amount"},
                {"text": "شماره پیگیری", "dataindex": "tracking_number"},
                {"text": "مغایرت گیری شده", "dataindex": "is_reconciled"}
            ]
            
            conn.close()
        except Exception as e:
            self.logger.error(f"خطا در دریافت تراکنش‌های پوز: {str(e)}")
            raise
    
    def get_reconciliation_results_data(self, bank_id):
        """دریافت نتایج مغایرت‌گیری با فیلترهای مشخص شده"""
        try:
            # ایجاد کوئری برای دریافت نتایج مغایرت‌گیری
            import sqlite3
            from config.settings import DB_PATH
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = """
                SELECT r.*, 
                       b1.bank_name as bank_name,
                       bt.amount as bank_amount,
                       bt.transaction_date as bank_date,
                       bt.transaction_type as bank_transaction_type,
                       at.transaction_amount as accounting_amount,
                       at.due_date as accounting_date,
                       at.transaction_type as accounting_transaction_type,
                       pt.transaction_amount as pos_amount,
                       pt.transaction_date as pos_date
                FROM ReconciliationResults r
                LEFT JOIN BankTransactions bt ON r.bank_record_id = bt.id
                LEFT JOIN AccountingTransactions at ON r.acc_id = at.id
                LEFT JOIN PosTransactions pt ON r.pos_id = pt.id
                LEFT JOIN Banks b1 ON bt.bank_id = b1.id
                WHERE 1=1
            """
            params = []
            
            if bank_id:
                query += " AND (bt.bank_id = ? OR at.bank_id = ? OR pt.bank_id = ?)"
                params.extend([bank_id, bank_id, bank_id])
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # تبدیل نتایج به لیست دیکشنری
            self.data = [dict(row) for row in rows]
            self.columns = [
                {"text": "شناسه", "dataindex": "id"},
                {"text": "بانک", "dataindex": "bank_name"},
                {"text": "شناسه رکورد بانک", "dataindex": "bank_record_id"},
                {"text": "مبلغ بانک", "dataindex": "bank_amount"},
                {"text": "تاریخ بانک", "dataindex": "bank_date"},
                {"text": "نوع تراکنش بانک", "dataindex": "bank_transaction_type"},
                {"text": "شناسه رکورد حسابداری", "dataindex": "acc_id"},
                {"text": "مبلغ حسابداری", "dataindex": "accounting_amount"},
                {"text": "تاریخ حسابداری", "dataindex": "accounting_date"},
                {"text": "نوع تراکنش حسابداری", "dataindex": "accounting_transaction_type"},
                {"text": "شناسه رکورد پوز", "dataindex": "pos_id"},
                {"text": "مبلغ پوز", "dataindex": "pos_amount"},
                {"text": "تاریخ پوز", "dataindex": "pos_date"},
                {"text": "شرح", "dataindex": "description"},
                {"text": "نوع تطبیق", "dataindex": "type_matched"},
                {"text": "تاریخ و زمان", "dataindex": "date_time"}
            ]
            
            conn.close()
        except Exception as e:
            self.logger.error(f"خطا در دریافت نتایج مغایرت‌گیری: {str(e)}")
            raise
    
    def display_data_in_table(self):
        """نمایش داده‌ها در جدول"""
        try:
            # پاک کردن جدول قبلی
            for widget in self.table_frame.winfo_children():
                widget.destroy()
                
            # تبدیل ستون‌ها به فرمت صحیح برای Tableview
            coldata = []
            dataindex_to_colindex = {}
            
            for i, col in enumerate(self.columns):
                # تنظیم عرض مناسب برای هر ستون با در نظر گرفتن متن فارسی
                width = 150  # عرض پیش‌فرض مناسب
                if col["dataindex"] in ["description", "customer_name"]:
                    width = 250  # عرض بیشتر برای ستون‌های توضیحات و نام مشتری
                elif col["dataindex"] in ["amount", "transaction_amount", "bank_amount", "accounting_amount", "pos_amount"]:
                    width = 140  # عرض مناسب برای ستون‌های مبلغ
                elif col["dataindex"] in ["id", "is_reconciled", "is_new_system"]:
                    width = 100  # عرض کمتر برای ستون‌های کوتاه
                elif col["dataindex"] in ["transaction_date", "due_date", "collection_date", "date_time", "bank_date", "accounting_date", "pos_date"]:
                    width = 120  # عرض مناسب برای تاریخ‌ها
                
                coldata.append({"text": col["text"], "stretch": True, "width": width})
                dataindex_to_colindex[col["dataindex"]] = i
            
            # تبدیل داده‌ها به فرمت مناسب برای Tableview (لیست تاپل‌ها)
            rowdata = []
            from utils.helpers import gregorian_to_persian
            
            for item in self.data:
                row = ["" for _ in range(len(self.columns))]
                for col in self.columns:
                    key = col["dataindex"]
                    col_index = dataindex_to_colindex[key]
                    if key in item:
                        # تبدیل مقادیر بولین به متن
                        if key == "is_reconciled":
                            row[col_index] = "بله" if item[key] == 1 else "خیر"
                        # تبدیل مقادیر is_new_system به متن
                        elif key == "is_new_system":
                            row[col_index] = "سیستم جدید" if item[key] == 1 else "سیستم قدیم"
                        # تبدیل تاریخ میلادی به شمسی
                        elif key in ["transaction_date", "due_date", "collection_date", "date_time", "bank_date", "accounting_date", "pos_date"] and item[key]:
                            try:
                                row[col_index] = gregorian_to_persian(str(item[key]))
                            except Exception as e:
                                self.logger.error(f"خطا در تبدیل تاریخ {item[key]}: {str(e)}")
                                row[col_index] = str(item[key])
                        # فرمت کردن مبالغ
                        elif key in ["amount", "transaction_amount", "bank_amount", "accounting_amount", "pos_amount"]:
                            try:
                                if item[key] is not None:
                                    # تبدیل به عدد و سپس فرمت‌بندی با جداکننده هزارگان
                                    amount_value = float(item[key])
                                    row[col_index] = f"{int(amount_value):,}"
                                else:
                                    row[col_index] = ""
                            except (ValueError, TypeError):
                                row[col_index] = str(item[key]) if item[key] is not None else ""
                        else:
                            # اطمینان از تبدیل صحیح به رشته و حذف مقادیر None
                            row[col_index] = str(item[key]) if item[key] is not None else ""
                rowdata.append(tuple(row))
            
            # ایجاد جدول با تنظیمات بهتر برای متن فارسی
            table = Tableview(
                master=self.table_frame,
                coldata=coldata,
                rowdata=rowdata,
                paginated=True,
                searchable=True,
                bootstyle="primary",
                stripecolor=("#f8f9fa", "#ffffff"),  # رنگ‌های بهتر برای خوانایی
                autofit=False,
                pagesize=12,  # تعداد مناسب برای نمایش بهتر
                height=18  # ارتفاع مناسب برای خوانایی
            )
            # تنظیم پکینگ جدول
            table.pack(fill="both", expand=True, padx=5, pady=5)
            
            # تنظیم فونت فارسی برای جدول با استفاده از روش غیرمستقیم
            try:
                # تلاش برای تنظیم فونت از طریق تغییر استایل
                style = ttk.Style()
                
                # تنظیم فونت برای جدول با فونت مناسب برای فارسی
                # استفاده از Tahoma به عنوان فونت پشتیبان چون بهتر با فارسی کار می‌کند
                table_font = ("Tahoma", 12)
                header_font = ("Tahoma", 13, "bold")
                
                style.configure("Treeview", font=table_font)
                style.configure("Treeview.Heading", font=header_font)
                
                # تنظیم ارتفاع سطرها برای نمایش بهتر متن فارسی
                style.configure("Treeview", rowheight=30)
                
                # تنظیم ترازبندی برای متن فارسی (راست به چپ)
                style.configure("Treeview.Heading", anchor="center")
                style.configure("Treeview", anchor="e")  # Right align for RTL text
                
            except Exception as e:
                self.logger.warning(f"تنظیم فونت فارسی با خطا مواجه شد: {str(e)}")
                # ادامه اجرا بدون توقف
            
            # ذخیره داده‌ها برای استفاده در صدور
            self.table_data = rowdata
        except Exception as e:
            self.logger.error(f"خطا در نمایش داده‌ها در جدول: {str(e)}")
            raise
    
    def export_to_excel(self):
        """صدور داده‌ها به فایل اکسل"""
        try:
            if not self.data:
                messagebox.showwarning("هشدار", "هیچ داده‌ای برای صدور وجود ندارد")
                return
            
            # دریافت مسیر ذخیره فایل
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="ذخیره فایل اکسل"
            )
            
            if not file_path:
                return
            
            # فیلتر کردن داده‌ها برای حذف کارمزدها از شیت اصلی
            filtered_data = []
            for row in self.table_data:
                # بررسی نوع تراکنش در ستون مربوطه
                transaction_type_index = None
                transaction_type_column_names = ["نوع تراکنش بانک", "نوع تراکنش", "transaction_type"]
                
                for i, col in enumerate(self.columns):
                    if col["text"] in transaction_type_column_names:
                        transaction_type_index = i
                        break
                
                # اگر ستون نوع تراکنش پیدا شد
                if transaction_type_index is not None:
                    transaction_type = row[transaction_type_index] if transaction_type_index < len(row) else None
                    # اگر نوع تراکنش کارمزد نیست، به لیست فیلتر شده اضافه کن
                    if transaction_type not in ["bank_fee", "BANK_FEES", "کارمزد بانکی"]:
                        filtered_data.append(list(row))  # تبدیل تاپل به لیست
                else:  # اگر ستون نوع تراکنش پیدا نشد، همه رکوردها را اضافه کن
                    filtered_data.append(list(row))  # تبدیل تاپل به لیست
                    
                # لاگ کردن برای اشکال‌زدایی
                self.logger.debug(f"نوع تراکنش: {row[transaction_type_index] if transaction_type_index is not None and transaction_type_index < len(row) else 'نامشخص'}, اضافه شده به لیست: {transaction_type_index is None or row[transaction_type_index] not in ['bank_fee', 'BANK_FEES', 'کارمزد بانکی']}")
            
            # اضافه کردن ستون اضافی بر اساس نوع گزارش
            column_names = [col["text"] for col in self.columns]
            
            # تعیین نوع ستون اضافی بر اساس نوع گزارش
            selected_table = self.selected_table_var.get()
            if selected_table == "بانک":
                additional_column_name = "acc_id"
            elif selected_table == "حسابداری":
                additional_column_name = "bank_rec_id"
            else:
                additional_column_name = None  # برای پوز و نتایج مغایرت گیری ستون اضافی نداریم
            
            # اضافه کردن نام ستون اضافی
            if additional_column_name:
                column_names.append(additional_column_name)
                # اضافه کردن ستون خالی به داده‌های فیلتر شده
                for row in filtered_data:
                    row.append("")  # ستون خالی
            
            df = pd.DataFrame(filtered_data, columns=column_names)
            
            # تنظیم فونت و استایل برای فایل اکسل
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            # ذخیره به فایل اکسل
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='گزارش')
                
                # تنظیم استایل‌ها
                workbook = writer.book
                worksheet = writer.sheets['گزارش']
                
                # تنظیم راست به چپ بودن کل شیت
                worksheet.sheet_view.rightToLeft = True
                
                # تنظیم استایل هدر
                header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                header_font = Font(name='Tahoma', size=12, bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # تنظیم فونت و راست به چپ بودن برای داده‌ها
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(name='Tahoma', size=11)
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # تنظیم عرض ستون‌ها
                for i, column in enumerate(worksheet.columns):
                    max_length = 0
                    column_letter = get_column_letter(i+1)
                    
                    # بررسی طول محتوای سلول‌ها
                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    # تنظیم عرض ستون با توجه به محتوا
                    adjusted_width = max(max_length + 4, 15)  # حداقل عرض 15 کاراکتر
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # اضافه کردن شیت کارمزدها
                self.add_bank_fees_sheet(writer)
            
            self.logger.info(f"داده‌ها با موفقیت به فایل {file_path} صادر شدند")
            self.status_var.set(f"داده‌ها با موفقیت به فایل اکسل صادر شدند")
            messagebox.showinfo("موفقیت", f"داده‌ها با موفقیت به فایل اکسل صادر شدند")
        except Exception as e:
            self.logger.error(f"خطا در صدور به اکسل: {str(e)}")
            self.status_var.set(f"خطا در صدور به اکسل: {str(e)}")
            messagebox.showerror("خطا", f"خطا در صدور به اکسل: {str(e)}")
    
    def add_bank_fees_sheet(self, writer):
        """اضافه کردن شیت کارمزدهای بانکی به فایل اکسل با جمع‌بندی روزانه"""
        try:
            import sqlite3
            from config.settings import DB_PATH
            
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # دریافت کارمزدهای بانکی
            query = """SELECT bt.*, b.bank_name 
                       FROM BankTransactions bt 
                       JOIN Banks b ON bt.bank_id = b.id 
                       WHERE bt.transaction_type IN ('bank_fee', 'BANK_FEES')"""
            
            cursor.execute(query)
            rows = cursor.fetchall()
            
            # اگر کارمزدی وجود نداشت، شیت اضافه نشود
            if not rows:
                self.logger.info("هیچ کارمزد بانکی یافت نشد")
                return
            
            # تبدیل نتایج به لیست دیکشنری
            fees_data = [dict(row) for row in rows]
            
            # تبدیل تاریخ‌ها به شمسی و گروه‌بندی بر اساس تاریخ
            from utils.helpers import gregorian_to_persian
            
            # گروه‌بندی کارمزدها بر اساس تاریخ
            daily_fees = {}
            for item in fees_data:
                if 'transaction_date' in item and item['transaction_date']:
                    try:
                        # تبدیل تاریخ به شمسی
                        persian_date = gregorian_to_persian(str(item['transaction_date']))
                        
                        # اضافه کردن به دیکشنری گروه‌بندی شده
                        if persian_date not in daily_fees:
                            daily_fees[persian_date] = {
                                'date': persian_date,
                                'total_amount': 0,
                                'bank_name': item['bank_name'],
                                'count': 0
                            }
                        
                        # جمع کردن مبلغ کارمزد
                        if 'amount' in item and item['amount'] is not None:
                            daily_fees[persian_date]['total_amount'] += float(item['amount'])
                            daily_fees[persian_date]['count'] += 1
                    except Exception as e:
                        self.logger.error(f"خطا در پردازش تاریخ: {str(e)}")
            
            # تبدیل دیکشنری به لیست برای ایجاد دیتافریم
            daily_fees_list = list(daily_fees.values())
            
            # اگر داده‌ای برای نمایش وجود نداشت
            if not daily_fees_list:
                self.logger.info("هیچ داده‌ای برای نمایش در شیت کارمزدها وجود ندارد")
                return
            
            # ستون‌های مورد نیاز برای نمایش
            columns = [
                {"text": "تاریخ", "dataindex": "date"},
                {"text": "بانک", "dataindex": "bank_name"},
                {"text": "جمع کارمزد (ریال)", "dataindex": "total_amount"},
                {"text": "تعداد تراکنش", "dataindex": "count"}
            ]
            
            # آماده‌سازی داده‌ها برای دیتافریم
            rows_data = []
            for item in daily_fees_list:
                row = []
                for col in columns:
                    key = col["dataindex"]
                    if key in item:
                        if key == "total_amount":
                            try:
                                if item[key] is not None:
                                    amount_value = float(item[key])
                                    row.append(f"{int(amount_value):,}")
                                else:
                                    row.append("")
                            except:
                                row.append(str(item[key]) if item[key] is not None else "")
                        else:
                            row.append(str(item[key]) if item[key] is not None else "")
                rows_data.append(row)
            
            # ایجاد دیتافریم
            column_names = [col["text"] for col in columns]
            df_fees = pd.DataFrame(rows_data, columns=column_names)
            
            # مرتب‌سازی بر اساس تاریخ
            df_fees = df_fees.sort_values(by="تاریخ")
            
            # اضافه کردن به اکسل
            df_fees.to_excel(writer, sheet_name='کارمزدها', index=False)
            
            # تنظیم استایل‌ها
            worksheet = writer.sheets['کارمزدها']
            
            # تنظیم راست به چپ بودن کل شیت
            worksheet.sheet_view.rightToLeft = True
            
            # تنظیم استایل هدر
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            header_font = Font(name='Tahoma', size=12, bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # تنظیم فونت و راست به چپ بودن برای داده‌ها
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name='Tahoma', size=11)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # تنظیم عرض ستون‌ها
            for i, column in enumerate(worksheet.columns):
                max_length = 0
                column_letter = get_column_letter(i+1)
                
                # بررسی طول محتوای سلول‌ها
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # تنظیم عرض ستون با توجه به محتوا
                adjusted_width = max(max_length + 4, 15)  # حداقل عرض 15 کاراکتر
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            self.logger.info("شیت کارمزدها با موفقیت اضافه شد")
            conn.close()
        except Exception as e:
            self.logger.error(f"خطا در اضافه کردن شیت کارمزدها: {str(e)}")
    
    def print_report(self):
        """چاپ گزارش"""
        try:
            if not self.data:
                messagebox.showwarning("هشدار", "هیچ داده‌ای برای چاپ وجود ندارد")
                return
            
            # ایجاد یک فایل HTML موقت برای چاپ
            import tempfile
            import webbrowser
            import os
            
            # ایجاد محتوای HTML
            # مسیر فونت
            font_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "fonts", "Vazir.ttf")
            font_path = font_path.replace('\\', '/')
            
            html_content = """<!DOCTYPE html>
            <html dir="rtl">
            <head>
                <meta charset="UTF-8">
                <title>گزارش</title>
                <style>
                    @font-face {
                        font-family: 'Vazir';
                        src: url('file:///{font_path}') format('truetype');
                    }
                    body { font-family: Vazir, Tahoma, Arial, sans-serif; direction: rtl; }
                    table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                    th, td { border: 1px solid #ddd; padding: 8px; text-align: right; }
                    th { background-color: #f2f2f2; }
                    h1, h2 { text-align: center; }
                    .report-header { margin-bottom: 20px; }
                    .report-footer { margin-top: 20px; text-align: center; }
                    @media print {
                        body { width: 21cm; height: 29.7cm; margin: 0; }
                        .no-print { display: none; }
                        button { display: none; }
                    }
                </style>
            </head>
            <body>
                <div class="report-header">
                    <h1>گزارش سیستم مغایرت‌گیری</h1>
                    <h2>{report_title}</h2>
                    <p>تاریخ گزارش: {report_date}</p>
                </div>
                
                <table>
                    <thead>
                        <tr>
                            {table_header}
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows}
                    </tbody>
                </table>
                
                <div class="report-footer">
                    <p>تعداد رکوردها: {record_count}</p>
                </div>
                
                <div class="no-print" style="text-align: center; margin-top: 20px;">
                    <button onclick="window.print()">چاپ</button>
                </div>
            </body>
            </html>
            """
            
            # ایجاد عنوان گزارش
            report_title = f"گزارش {self.selected_table_var.get()}"
            if self.selected_bank_var.get() != "همه موارد":
                report_title += f" - بانک {self.selected_bank_var.get()}"
            
            # ایجاد تاریخ گزارش
            from datetime import datetime
            import jdatetime
            gregorian_date = datetime.now()
            jalali_date = jdatetime.datetime.fromgregorian(datetime=gregorian_date)
            report_date = jalali_date.strftime("%Y/%m/%d %H:%M:%S")
            
            # ایجاد هدر جدول
            table_header = ""
            for col in self.columns:
                table_header += f"<th>{col['text']}</th>\n"
            
            # ایجاد ردیف‌های جدول
            table_rows = ""
            
            for row in self.table_data:
                table_rows += "<tr>\n"
                for i, value in enumerate(row):
                    table_rows += f"<td>{value}</td>\n"
                table_rows += "</tr>\n"
            
            # تکمیل محتوای HTML با استفاده از f-string برای جلوگیری از خطای فرمت‌بندی
            html_content = html_content.format(font_path=font_path, report_title=report_title, report_date=report_date, table_header=table_header, table_rows=table_rows, record_count=len(self.table_data))
            
            # ایجاد فایل موقت
            with tempfile.NamedTemporaryFile(delete=False, suffix='.html', mode='w', encoding='utf-8') as f:
                f.write(html_content)
                temp_file_path = f.name
            
            # باز کردن فایل در مرورگر برای چاپ
            webbrowser.open('file://' + os.path.realpath(temp_file_path))
            
            self.logger.info("فایل چاپ با موفقیت ایجاد شد")
            self.status_var.set("فایل چاپ با موفقیت ایجاد شد")
        except Exception as e:
            self.logger.error(f"خطا در چاپ گزارش: {str(e)}")
            self.status_var.set(f"خطا در چاپ گزارش: {str(e)}")
            messagebox.showerror("خطا", f"خطا در چاپ گزارش: {str(e)}")
    
    def export_to_pdf(self):
        """صدور گزارش به فایل PDF"""
        try:
            if not self.data:
                messagebox.showwarning("هشدار", "هیچ داده‌ای برای صدور وجود ندارد")
                return
            
            # دریافت مسیر ذخیره فایل
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                title="ذخیره فایل PDF"
            )
            
            if not file_path:
                return
            
            # استفاده از کتابخانه reportlab برای ایجاد PDF
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.lib.enums import TA_RIGHT, TA_CENTER
                from reportlab.pdfgen import canvas
                import os
                
                # Import for RTL text handling
                try:
                    from bidi.algorithm import get_display
                    from arabic_reshaper import reshape
                    bidi_available = True
                except ImportError:
                    bidi_available = False
                    self.logger.warning("کتابخانه‌های bidi و arabic-reshaper یافت نشد. متن فارسی ممکن است به درستی نمایش داده نشود")
                
                def format_persian_text(text):
                    """تنظیم متن فارسی برای نمایش صحیح در PDF"""
                    if not text or not isinstance(text, str):
                        return str(text) if text is not None else ""
                    
                    if bidi_available:
                        try:
                            # Reshape Arabic/Persian characters
                            reshaped_text = reshape(text)
                            # Apply bidirectional algorithm
                            display_text = get_display(reshaped_text)
                            return display_text
                        except Exception as e:
                            self.logger.warning(f"خطا در تنظیم متن فارسی: {str(e)}")
                            return text
                    else:
                        # Fallback: at least reverse the string for basic RTL support
                        try:
                            # Simple reversal for Persian text (not perfect but better than nothing)
                            import re
                            if re.search(r'[\u0600-\u06FF]', text):
                                return text[::-1]  # Simple reversal for Persian characters
                            return text
                        except:
                            return text
                
                # ثبت فونت فارسی
                font_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "assets", "fonts", "Vazir.ttf")
                # اطمینان از وجود فایل فونت
                if not os.path.exists(font_path):
                    self.logger.error(f"فایل فونت در مسیر {font_path} یافت نشد")
                    messagebox.showerror("خطا", f"فایل فونت در مسیر {font_path} یافت نشد")
                    return
                self.logger.info(f"فایل فونت در مسیر {font_path} یافت شد")
                pdfmetrics.registerFont(TTFont('Vazir', font_path))
                
                # ایجاد استایل‌های متن
                styles = getSampleStyleSheet()
                styles.add(ParagraphStyle(name='RTL', fontName='Vazir', alignment=TA_RIGHT))
                
                # ایجاد داکیومنت PDF
                doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
                elements = []
                
                # عنوان گزارش
                title_style = ParagraphStyle(
                    'Title',
                    parent=styles['Title'],
                    fontName='Vazir',
                    alignment=TA_RIGHT,
                    fontSize=16
                )
                
                report_title = f"گزارش {self.selected_table_var.get()}"
                if self.selected_bank_var.get() != "همه موارد":
                    report_title += f" - بانک {self.selected_bank_var.get()}"
                
                formatted_title = format_persian_text(report_title)
                elements.append(Paragraph(formatted_title, title_style))
                elements.append(Spacer(1, 20))
                
                # تاریخ گزارش
                from datetime import datetime
                import jdatetime
                date_style = ParagraphStyle(
                    'Date',
                    parent=styles['Normal'],
                    fontName='Vazir',
                    alignment=TA_RIGHT,
                    fontSize=10
                )
                gregorian_date = datetime.now()
                jalali_date = jdatetime.datetime.fromgregorian(datetime=gregorian_date)
                report_date = jalali_date.strftime("%Y/%m/%d %H:%M:%S")
                formatted_date_text = format_persian_text(f"تاریخ گزارش: {report_date}")
                elements.append(Paragraph(formatted_date_text, date_style))
                elements.append(Spacer(1, 20))
                
                # ایجاد داده‌های جدول
                table_data = []
                
                # هدر جدول
                header_row = [format_persian_text(col['text']) for col in self.columns]
                table_data.append(header_row)
                
                # ردیف‌های جدول
                for row in self.table_data:
                    data_row = []
                    # چون self.table_data یک لیست از تاپل‌هاست، مستقیماً از عناصر آن استفاده می‌کنیم
                    for value in row:
                        formatted_value = format_persian_text(str(value) if value is not None else "")
                        data_row.append(formatted_value)
                    table_data.append(data_row)
                
                # ایجاد جدول
                table = Table(table_data, repeatRows=1)
                
                # استایل جدول
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Vazir'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elements.append(table)
                
                # پاورقی
                elements.append(Spacer(1, 20))
                footer_style = ParagraphStyle(
                    'Footer',
                    parent=styles['Normal'],
                    fontName='Vazir',
                    alignment=TA_RIGHT,
                    fontSize=10
                )
                formatted_footer_text = format_persian_text(f"تعداد رکوردها: {len(self.table_data)}")
                elements.append(Paragraph(formatted_footer_text, footer_style))
                
                # ساخت PDF
                doc.build(elements)
                
                self.logger.info(f"گزارش با موفقیت به فایل {file_path} صادر شد")
                self.status_var.set("گزارش با موفقیت به PDF صادر شد")
                messagebox.showinfo("موفقیت", "گزارش با موفقیت به PDF صادر شد")
            except ImportError:
                # اگر کتابخانه reportlab نصب نشده باشد
                self.logger.error("کتابخانه‌های مورد نیاز برای PDF نصب نشده است")
                self.status_var.set("خطا: کتابخانه‌های مورد نیاز PDF نصب نشده است")
                error_msg = ("برای صدور به PDF با پشتیبانی فارسی نیاز به نصب کتابخانه‌های زیر دارید:\n\n"
                             "pip install -r requirements_pdf_persian.txt\n\n"
                             "یا:\n\n"
                             "pip install reportlab python-bidi arabic-reshaper jdatetime")
                messagebox.showerror("خطا", error_msg)
        except Exception as e:
            self.logger.error(f"خطا در صدور به PDF: {str(e)}")
            self.status_var.set(f"خطا در صدور به PDF: {str(e)}")
            messagebox.showerror("خطا", f"خطا در صدور به PDF: {str(e)}")