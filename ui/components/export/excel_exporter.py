"""
Excel Export Module
جدا شده از report_tab.py برای ماژولار کردن کد
"""
import os
import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import sqlite3
import logging
from config.settings import DB_PATH
from utils.helpers import gregorian_to_persian


class ExcelExporter:
    """کلاس صدور اطلاعات به فایل اکسل"""
    
    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger(__name__)
    
    def export_to_excel(self, data, columns, selected_table, table_data):
        """صدور داده‌ها به فایل اکسل"""
        try:
            if not data:
                messagebox.showwarning("هشدار", "هیچ داده‌ای برای صدور وجود ندارد")
                return False
            
            # دریافت مسیر ذخیره فایل
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="ذخیره فایل اکسل"
            )
            
            if not file_path:
                return False
            
            # فیلتر کردن داده‌ها برای حذف کارمزدها از شیت اصلی
            filtered_data = self._filter_bank_fees(table_data, columns)
            
            # اضافه کردن ستون اضافی بر اساس نوع گزارش
            column_names = [col["text"] for col in columns]
            self._add_additional_column(filtered_data, column_names, selected_table)
            
            # تبدیل داده‌های فیلتر شده به دیتافریم پانداس
            df = pd.DataFrame(filtered_data, columns=column_names)
            
            # ذخیره به فایل اکسل
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='گزارش')
                
                # تنظیم استایل‌ها
                self._apply_excel_styles(writer)
                
                # اضافه کردن شیت کارمزدها
                self._add_bank_fees_sheet(writer)
            
            self.logger.info(f"داده‌ها با موفقیت به فایل {file_path} صادر شدند")
            messagebox.showinfo("موفقیت", "داده‌ها با موفقیت به فایل اکسل صادر شدند")
            return True
            
        except Exception as e:
            error_msg = f"خطا در صدور به اکسل: {str(e)}"
            self.logger.error(error_msg)
            messagebox.showerror("خطا", error_msg)
            return False
    
    def _filter_bank_fees(self, table_data, columns):
        """فیلتر کردن کارمزدهای بانکی از داده‌ها"""
        filtered_data = []
        
        for row in table_data:
            # بررسی نوع تراکنش در ستون مربوطه
            transaction_type_index = None
            transaction_type_column_names = ["نوع تراکنش بانک", "نوع تراکنش", "transaction_type"]
            
            for i, col in enumerate(columns):
                if col["text"] in transaction_type_column_names:
                    transaction_type_index = i
                    break
            
            # اگر ستون نوع تراکنش پیدا شد
            if transaction_type_index is not None:
                transaction_type = row[transaction_type_index] if transaction_type_index < len(row) else None
                # اگر نوع تراکنش کارمزد نیست، به لیست فیلتر شده اضافه کن
                if transaction_type not in ["bank_fee", "BANK_FEES", "کارمزد بانکی"]:
                    filtered_data.append(list(row))  # تبدیل تاپل به لیست
            else:  # اگر ستون نوع تراکنش پیدا نشد، همه رکوردها را اضافه کن
                filtered_data.append(list(row))  # تبدیل تاپل به لیست
        
        return filtered_data
    
    def _add_additional_column(self, filtered_data, column_names, selected_table):
        """اضافه کردن ستون اضافی بر اساس نوع گزارش"""
        # تعیین نوع ستون اضافی بر اساس نوع گزارش
        if selected_table == "بانک":
            additional_column_name = "acc_id"
        elif selected_table == "حسابداری":
            additional_column_name = "bank_rec_id"
        else:
            additional_column_name = None  # برای پوز و نتایج مغایرت گیری ستون اضافی نداریم
        
        # اضافه کردن نام ستون اضافی
        if additional_column_name:
            column_names.append(additional_column_name)
            # اضافه کردن ستون خالی به داده‌های فیلتر شده
            for row in filtered_data:
                row.append("")  # ستون خالی
    
    def _apply_excel_styles(self, writer):
        """اعمال استایل‌ها به فایل اکسل"""
        workbook = writer.book
        worksheet = writer.sheets['گزارش']
        
        # تنظیم راست به چپ بودن کل شیت
        worksheet.sheet_view.rightToLeft = True
        
        # تنظیم استایل هدر
        header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        header_font = Font(name='Tahoma', size=12, bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تنظیم فونت و راست به چپ بودن برای داده‌ها
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Tahoma', size=11)
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # تنظیم عرض ستون‌ها
        for i, column in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(i+1)
            
            # بررسی طول محتوای سلول‌ها
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # تنظیم عرض ستون با توجه به محتوا
            adjusted_width = max(max_length + 4, 15)  # حداقل عرض 15 کاراکتر
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_bank_fees_sheet(self, writer):
        """اضافه کردن شیت کارمزدهای بانکی به فایل اکسل"""
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # دریافت کارمزدهای بانکی
            query = """SELECT bt.*, b.bank_name 
                       FROM BankTransactions bt 
                       JOIN Banks b ON bt.bank_id = b.id 
                       WHERE bt.transaction_type IN ('bank_fee', 'BANK_FEES')"""
            
            cursor.execute(query)
            rows = cursor.fetchall()
            
            # اگر کارمزدی وجود نداشت، شیت اضافه نشود
            if not rows:
                self.logger.info("هیچ کارمزد بانکی یافت نشد")
                return
            
            # تبدیل نتایج به لیست دیکشنری
            fees_data = [dict(row) for row in rows]
            
            # گروه‌بندی کارمزدها بر اساس تاریخ
            daily_fees = self._group_fees_by_date(fees_data)
            
            if not daily_fees:
                self.logger.info("هیچ داده‌ای برای نمایش در شیت کارمزدها وجود ندارد")
                return
            
            # ایجاد دیتافریم و اضافه به اکسل
            self._create_fees_dataframe(writer, daily_fees)
            
            self.logger.info("شیت کارمزدها با موفقیت اضافه شد")
            conn.close()
            
        except Exception as e:
            self.logger.error(f"خطا در اضافه کردن شیت کارمزدها: {str(e)}")
    
    def _group_fees_by_date(self, fees_data):
        """گروه‌بندی کارمزدها بر اساس تاریخ"""
        daily_fees = {}
        
        for item in fees_data:
            if 'transaction_date' in item and item['transaction_date']:
                try:
                    # تبدیل تاریخ به شمسی
                    persian_date = gregorian_to_persian(str(item['transaction_date']))
                    
                    # اضافه کردن به دیکشنری گروه‌بندی شده
                    if persian_date not in daily_fees:
                        daily_fees[persian_date] = {
                            'date': persian_date,
                            'total_amount': 0,
                            'bank_name': item['bank_name'],
                            'count': 0
                        }
                    
                    # جمع کردن مبلغ کارمزد
                    if 'amount' in item and item['amount'] is not None:
                        daily_fees[persian_date]['total_amount'] += float(item['amount'])
                        daily_fees[persian_date]['count'] += 1
                        
                except Exception as e:
                    self.logger.error(f"خطا در پردازش تاریخ: {str(e)}")
        
        return list(daily_fees.values())
    
    def _create_fees_dataframe(self, writer, daily_fees_list):
        """ایجاد دیتافریم برای شیت کارمزدها"""
        # ستون‌های مورد نیاز برای نمایش
        columns = [
            {"text": "تاریخ", "dataindex": "date"},
            {"text": "بانک", "dataindex": "bank_name"}, 
            {"text": "جمع کارمزد (ریال)", "dataindex": "total_amount"},
            {"text": "تعداد تراکنش", "dataindex": "count"}
        ]
        
        # آماده‌سازی داده‌ها برای دیتافریم
        rows_data = []
        for item in daily_fees_list:
            row = []
            for col in columns:
                key = col["dataindex"]
                if key in item:
                    if key == "total_amount":
                        try:
                            if item[key] is not None:
                                amount_value = float(item[key])
                                row.append(f"{int(amount_value):,}")
                            else:
                                row.append("")
                        except:
                            row.append(str(item[key]) if item[key] is not None else "")
                    else:
                        row.append(str(item[key]) if item[key] is not None else "")
            rows_data.append(row)
        
        # ایجاد دیتافریم
        column_names = [col["text"] for col in columns]
        df_fees = pd.DataFrame(rows_data, columns=column_names)
        
        # مرتب‌سازی بر اساس تاریخ
        df_fees = df_fees.sort_values(by="تاریخ")
        
        # اضافه کردن به اکسل
        df_fees.to_excel(writer, sheet_name='کارمزدها', index=False)
        
        # تنظیم استایل‌ها برای شیت کارمزدها
        self._apply_fees_sheet_styles(writer)
    
    def _apply_fees_sheet_styles(self, writer):
        """اعمال استایل‌ها به شیت کارمزدها"""
        worksheet = writer.sheets['کارمزدها']
        
        # تنظیم راست به چپ بودن کل شیت
        worksheet.sheet_view.rightToLeft = True
        
        # تنظیم استایل هدر
        header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        header_font = Font(name='Tahoma', size=12, bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تنظیم فونت و راست به چپ بودن برای داده‌ها
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Tahoma', size=11)
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # تنظیم عرض ستون‌ها
        for i, column in enumerate(worksheet.columns):
            max_length = 0
            column_letter = get_column_letter(i+1)
            
            # بررسی طول محتوای سلول‌ها
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # تنظیم عرض ستون با توجه به محتوا
            adjusted_width = max(max_length + 4, 15)  # حداقل عرض 15 کاراکتر
            worksheet.column_dimensions[column_letter].width = adjusted_width
